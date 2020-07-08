import openpyxl


class HomepageData:

    @staticmethod
    def gettestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C://Users//Fleek//Desktop//Demo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Testcase2":
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
